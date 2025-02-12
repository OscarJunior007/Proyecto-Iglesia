

# import openpyxl module 
import openpyxl 
dic = {
    "numero_celular": "Celular",
    "direccion": "Direccion",
    "cedula_user": "Numero identifacion del lider",
    "email": "Email",
    "nombre":"Nombre del usuario",
    "nombre_celula":"Nombre de la celula",
    "dia_celula":"Dia de la cedula",
    "name_lider":"Nombre del lider",
    "name_anfitrion":"Nombre del anfitrion",
    "name_sublider":"Nombre del sublider",
    "asistencias_celula":"Total asistencia",
    "fecha_reunion":"Fecha de la reunion",
    "asisten_iglesia":"Asistentes a la iglesia",
    "new_people":"Nuevos en la celula"
}

list = ["celulaId","id","profile","lider_cedula"]

def remove(data):
    for i in data:
        for j in list:
            del(i[j])



def export(data, file_name):
    # Crear un nuevo libro de Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # Añadir los encabezados de columna
    if data:
        remove(data)
        headers = data[0].keys()
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=dic.get(header))
        
        # Añadir los datos del diccionario a las celdas correspondientes
        for row, record in enumerate(data, start=2):
            for col, (header, value) in enumerate(record.items(), start=1):
                sheet.cell(row=row, column=col, value=value)
    
    # Guardar el archivo de Excel
    workbook.save(file_name)